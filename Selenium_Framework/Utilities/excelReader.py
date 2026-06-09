import openpyxl


def get_data(path, sheetname):

    workbook = openpyxl.load_workbook(path)

    sheet = workbook[sheetname]

    data = []

    for i in range(2, sheet.max_row + 1):

        row = []

        for j in range(1, sheet.max_column + 1):

            row.append(
                sheet.cell(i, j).value
            )

        data.append(tuple(row))

    workbook.close()

    return data